import openpyxl as pyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
import pandas as pd
import os


def create_excel_from_csv(input_file_path, output_path, output_name='output'):
    """ create_excel_from_csv(input_file_path, output_path, output_name)
    input_file_path = full input file path (including extension)
    output_path = output folder path
    output_name = output file name (without extension)
    Given a csv file location, will create an Excel file at the desired output location."""
    file = pd.read_csv(input_file_path)
    path = create_path(output_path, output_name, '.xlsx')
    writer = pd.ExcelWriter(path)
    file.to_excel(writer, sheet_name='Data', index=False)
    writer.save()
    return path, file.shape[1], file.shape[0]


def create_excel_from_df(df, output_path, output_name):
    """ create_excel_from_df(df, output_path, output_name)
    df = DataFrame containing the data
    output_path = output folder path
    output_name = output file name (without extension)
    Given a DataFrame, will create an Excel file at the desired output location."""
    path = create_path(output_path, output_name, '.xlsx')
    writer = pd.ExcelWriter(path)
    df.to_excel(writer, sheet_name='Data', index=False)
    writer.save()
    return path, df.shape[1], df.shape[0]


def format_excel(file_path, x_max, y_max):
    """ format_excel(file_path, x_max, y_max)
    file_path = the full file path
    x_max, y_max = the dimensions of the file
    Given an excel file and its dimensions, this function formats the sheet following the expected BT output form.
    RETURNS None"""
    wb = pyxl.load_workbook(file_path)
    ws = wb['Data']

    ws.freeze_panes = 'A2'
    ws.insert_cols(2)
    formula = '=((A2:A' + str(y_max + 1) + '- 14400000)/86400000) + DATE(1970,1,1)'
    ws['B2'] = formula
    ws.formula_attributes['B2'] = {'t': 'array', 'ref': 'B2:B' + str(y_max + 1)}
    wb.save(file_path)
    wb = pyxl.load_workbook(file_path)
    ws = wb['Data']
    col = ws.column_dimensions['B']
    col.number_format = 'h:mm:ss AM/PM'
    ws['B2'].number_format = 'h:mm:ss AM/PM'
    ws['B1'].number_format = 'General'
    ws['B1'].font = Font(bold=True)
    ws['B1'] = 'Time'
    ws['A1'] = 'Datetime'

    rule = ColorScaleRule(start_type='min', start_color='6589C1',
                          mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                          end_type='max', end_color='E5726F')
    ws.conditional_formatting.add('C2:' + excel_column_name(x_max + 1) + str(y_max + 1), rule)

    for i in range(3, x_max + 2):
        column = excel_column_name(i)
        ws[column + str(y_max + 2)] = '=SUM(' + column + '2:' + column + str(y_max + 1) + ')'
        ws[column + str(y_max + 2)].font = Font(bold=True)
    ws['A' + str(y_max + 2)] = 'Sum per movement'
    ws['A' + str(y_max + 2)].font = Font(bold=True)

    wb.save(file_path)


def excel_column_name(n):
    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA.
        By 'devon' on StackOverflow https://stackoverflow.com/questions/7261936/convert-an-excel-or-spreadsheet-
        column-letter-to-its-number-in-pythonic-fashion"""
    name = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(r + ord('A')) + name
    return name


def create_path(path, name, ext):
    """ create_path(path, name, ext)
    path = folder path of the file location
    name = name of the file
    ext = extension of the file
    Creates the complete file path, based upon the OS (to use correct slashes).
    RETURNS string"""
    if os.name == 'nt':
        return path + '\\' + name + ext
    else:
        return path + '/' + name + ext